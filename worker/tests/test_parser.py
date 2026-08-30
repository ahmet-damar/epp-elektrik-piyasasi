"""EPP — Parser testleri (worker/parser.py).

Bu sentetik xlsx, 2026 Ocak EPDK Elektrik Piyasası Sektör Raporu Ek'i
(gerçek dosya) ile doğrulanmış gerçek yapıyı birebir taklit eder — bkz.
worker/parser.py modül notu (İLLER/İL/İl Adı başlık varyantları, T2/T3/T5/T6/
T7/T9/T10'un uzun-format/tek-boyutlu gerçek yapısı, aynı kanonik kaynağa
eşlenen birden fazla ham sütun, İstanbul'un iki dağıtım bölgesine bölünmesi).
Uçtan uca doğruluğu worker/tests/golden/ ile aynı Eskişehir 202601
değerleriyle worker/kpi.py üzerinden çapraz kontrol ediyoruz.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from worker import kpi, parser

GOLDEN_BEKLENEN = Path(__file__).parent / "golden" / "expected" / "kpi_expected.json"


def _sentetik_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- T1: kurulu güç (il×kaynak matrisi, "İLLER" başlığı) ---
    t1 = wb.create_sheet("Tablo 1")
    t1.append(
        [
            "Tablo 1 - Lisanslı Elektrik Kurulu Gücünün İl ve Kaynak Bazında Dağılımı (MW)"
        ]
    )
    t1.append(
        [
            "İLLER",
            "Akarsu",
            "Barajlı",
            "Doğal Gaz",
            "Rüzgar",
            "Güneş",
            "Linyit",
            "Genel Toplam",
        ]
    )
    # Eskişehir: Akarsu+Barajlı ayrı sütun (Hidrolik'e toplanmalı: 80+120=200)
    t1.append(["Eskişehir", 80.0, 120.0, 900.0, 300.0, 150.0, 450.0, 2000.0])
    t1.append(["TÜRKİYE", 80.0, 120.0, 900.0, 300.0, 150.0, 450.0, 2000.0])

    # --- T4: lisanssız kurulu güç ---
    t4 = wb.create_sheet("Tablo 4")
    t4.append(
        [
            "Tablo 4 - Lisanssız Elektrik Kurulu Gücünün İl ve Kaynak Bazında Dağılımı (MW)"
        ]
    )
    t4.append(["İLLER", "Güneş", "Rüzgar"])
    t4.append(["Eskişehir", 50.0, 10.0])

    # --- T2: üretim, ülke geneli kaynak bazında (uzun tek-boyutlu, "Doğalgaz" yazımı) ---
    t2 = wb.create_sheet("Tablo 2")
    t2.append(
        ["Tablo 2 - Lisanslı Elektrik Üretiminin Kaynak Bazında Aylık Gelişimi (MWh)"]
    )
    t2.append(["Kaynak Türü", "OCAK"])
    t2.append(["Doğalgaz", 400000.0])  # "Doğal Gaz" değil, boşluksuz yazım
    t2.append(["Rüzgar", 120000.0])
    t2.append(["Güneş", 30000.0])
    t2.append(["Hidrolik", 80000.0])
    t2.append(["Linyit", 250000.0])
    t2.append(["TOPLAM", 880000.0])

    # --- T3: üretim, il bazında toplam (kaynak kırılımı yok) ---
    t3 = wb.create_sheet("Tablo 3")
    t3.append(
        ["Tablo 3 - Lisanslı Elektrik Üretiminin İl Bazında Aylık Gelişimi (MWh)"]
    )
    t3.append(["İLLER", "OCAK"])
    t3.append(["ESKİŞEHİR", 880000.0])
    t3.append(["TOPLAM", 880000.0])

    # --- T5: lisanssız üretim, kaynak bazında ---
    t5 = wb.create_sheet("Tablo 5")
    t5.append(
        [
            "Tablo 5 - Brüt Lisanssız Elektrik Üretiminin Kaynak Bazında Aylık Gelişimi (MWh)"
        ]
    )
    t5.append(["Kaynak Türü", "2026 OCAK"])
    t5.append(["Güneş ", 5000.0])  # sondaki boşluk
    t5.append(["Rüzgar", 1000.0])
    t5.append(["Genel Toplam", 6000.0])

    # --- T6: lisanssız üretim, il bazında toplam ---
    t6 = wb.create_sheet("Tablo 6")
    t6.append(
        ["Tablo 6 - Brüt Lisanssız Elektrik Üretiminin İl Bazında Aylık Gelişimi (MWh)"]
    )
    t6.append(["İLLER", "OCAK"])
    t6.append(["ESKİŞEHİR", 6000.0])

    # --- T7: faturalanan tüketim, ülke geneli tür bazında (uzun format, mutabakat) ---
    t7 = wb.create_sheet("Tablo 7")
    t7.append(
        [
            "Tablo 7 - Faturalanan Elektrik Tüketiminin Tüketici Türü Bazında Aylık Gelişimi (MWh)"
        ]
    )
    t7.append(["", "", "2026\nOcak"])
    t7.append(["İl Adı", "Tüketici Grubu", "Miktar"])
    t7.append(["TÜRKİYE", "Aydınlatma", 5000.0])
    t7.append([None, "Kamu ve Özel Hizmetler Sektörü ile Diğer", 60000.0])
    t7.append([None, "Mesken", 120000.0])
    t7.append([None, "Sanayi", 240000.0])
    t7.append([None, "Tarımsal Faaliyetler", 20000.0])
    t7.append([None, "Toplam", 445000.0])

    # --- T9: tüketici sayısı, ülke geneli tür bazında (uzun format, mutabakat) ---
    t9 = wb.create_sheet("Tablo 9")
    t9.append(
        [
            "Tablo 9 - Elektrik Tüketici Sayısının Tüketici Türü Bazında Aylık Gelişimi (Adet)"
        ]
    )
    t9.append(["", "", "2026\nOcak"])
    t9.append(["", "Tüketici Grubu", "Sayı"])
    t9.append(["TÜRKİYE", "Aydınlatma", 300])
    t9.append([None, "Kamu ve Özel Hizmetler Sektörü ile Diğer", 4000])
    t9.append([None, "Mesken", 250000])
    t9.append([None, "Sanayi", 1200])
    t9.append([None, "Tarımsal Faaliyetler", 800])
    t9.append([None, "Toplam", 256300])

    # --- T10: tüketici sayısı, il bazında (uzun format; İstanbul-tarzı bölünme dahil) ---
    t10 = wb.create_sheet("Tablo 10")
    t10.append(
        ["Tablo 10 - Elektrik Tüketici Sayısının İl Türü Bazında Aylık Gelişimi (Adet)"]
    )
    t10.append(["", "", "2026\nOcak"])
    t10.append(["İl Adı", "Tüketici Grubu", "Sayı"])
    t10.append(["ESKİŞEHİR", "Aydınlatma", 300])
    t10.append([None, "Kamu ve Özel Hizmetler Sektörü ile Diğer", 4000])
    t10.append([None, "Mesken", 250000])
    t10.append([None, "Sanayi", 1200])
    t10.append([None, "Tarımsal Faaliyetler", 800])
    t10.append([None, "İl Toplam", 256300])
    # İstanbul iki dağıtım bölgesine bölünmüş — aynı il_kodu'na toplanmalı
    t10.append(["İstanbul (Anadolu)", "Mesken", 100])
    t10.append([None, "İl Toplam", 100])
    t10.append(["İstanbul (Avrupa)", "Mesken", 50])
    t10.append([None, "İl Toplam", 50])

    # --- T11: tüketim, il×grup matrisi (P0-2, "İL" başlığı — "İLLER" değil) ---
    t11 = wb.create_sheet("Tablo 11")
    t11.append(
        [
            "Tablo 11 - Faturalanan Elektrik Tüketiminin İl ve Tüketici Türü Bazında Dağılımı (iletim-dağıtım kırılımlı) (MWh)"
        ]
    )
    t11.append(
        [
            "İL",
            "Aydınlatma",
            "Kamu ve Özel Hizmetler Sektörü ile Diğer",
            "Mesken",
            "Sanayi-DAĞITIM",
            "Sanayi-İLETİM",
            "Tarımsal Faaliyetler",
        ]
    )
    t11.append(
        [
            "ESKİŞEHİR",
            "5.000,00",
            "60.000,00",
            "120.000,00",
            "90.000,00",
            "150.000,00",
            "20.000,00",
        ]
    )

    # --- T13: serbest tüketici (il × tur × grup, iki paralel değer bloğu) ---
    t13 = wb.create_sheet("Tablo 13")
    t13.append(
        [
            "Tablo 13- İl Bazında Serbest Tüketici Bilgileri (Tüketim Miktarı - Tüketici Sayısı)"
        ]
    )
    t13.append(
        [
            "İl Adı",
            "Elektrik Tüketici Türü",
            "Tüketim Miktarı(MWh)",
            None,
            None,
            None,
            None,
            None,
            "Tüketici Sayısı",
            None,
            None,
            None,
            None,
            None,
        ]
    )
    t13.append(
        [
            None,
            None,
            "Mesken",
            "Sanayi",
            "Kamu/Özel ",
            "Tarımsal",
            "Aydınlatma",
            "Toplam",
            "Mesken",
            "Sanayi",
            "Kamu/Özel ",
            "Tarımsal",
            "Aydınlatma",
            "Toplam",
        ]
    )
    # Eskişehir: 3 tur satırı + atlanacak bir "İl Toplam" satırı; Tarımsal
    # sütunu ilk turda boş (NULL, 0 DEĞİL), üçüncü turda Mesken negatif
    # (dogrula_serbest_tuketici'nin red yolu için).
    t13.append(
        [
            "ESKİŞEHİR",
            "Serbest Tüketici",
            100.0,
            200.0,
            50.0,
            None,
            10.0,
            360.0,
            5,
            10,
            3,
            None,
            1,
            19,
        ]
    )
    t13.append(
        [
            None,
            "ST Olma Hakkı Bulunmayan Aboneler",
            20.0,
            30.0,
            5.0,
            1.0,
            2.0,
            58.0,
            100,
            50,
            20,
            5,
            2,
            177,
        ]
    )
    t13.append(
        [
            None,
            "ST Olma Hakkını Kullanmayan Aboneler",
            -5.0,
            0.0,
            0.0,
            0.0,
            0.0,
            -5.0,
            1,
            0,
            0,
            0,
            0,
            1,
        ]
    )
    t13.append(
        [
            None,
            "İl Toplam",
            115.0,
            230.0,
            55.0,
            1.0,
            12.0,
            413.0,
            106,
            60,
            23,
            5,
            3,
            197,
        ]
    )
    # Adana: ileri-doldurma tek satırlık bir tur ile de doğrulanır.
    t13.append(
        [
            "ADANA",
            "Serbest Tüketici",
            1.0,
            2.0,
            3.0,
            4.0,
            5.0,
            15.0,
            10,
            20,
            30,
            40,
            50,
            150,
        ]
    )
    t13.append(
        [None, "İl Toplam", 1.0, 2.0, 3.0, 4.0, 5.0, 15.0, 10, 20, 30, 40, 50, 150]
    )
    # İstanbul: Anadolu + Avrupa aynı il_kodu'na (34) düşer, toplanmalı -
    # 2026-08-31 canlı veri incelemesinde bulunan silent-drop hatası
    # (ON CONFLICT DO NOTHING bir tarafı sessizce eliyordu) için regresyon.
    t13.append(
        [
            "İSTANBUL (ANADOLU)",
            "Serbest Tüketici",
            10.0,
            20.0,
            30.0,
            40.0,
            50.0,
            150.0,
            1,
            2,
            3,
            4,
            5,
            15,
        ]
    )
    t13.append(
        [
            "İSTANBUL (AVRUPA)",
            "Serbest Tüketici",
            100.0,
            200.0,
            300.0,
            400.0,
            500.0,
            1500.0,
            10,
            20,
            30,
            40,
            50,
            150,
        ]
    )

    return wb


@pytest.fixture(scope="module")
def wb() -> openpyxl.Workbook:
    return _sentetik_workbook()


def test_tablo11_p0_2_ayri_satir(wb: openpyxl.Workbook) -> None:
    df = parser.tablo11_tuketim_oku(wb["Tablo 11"], 202601, "Tablo 11")
    assert len(df) == 6
    sanayi = df[df["grup"] == "Sanayi"]
    assert set(sanayi["baglanti"]) == {"iletim", "dagitim"}
    iletim = sanayi.loc[sanayi["baglanti"] == "iletim", "tuketim_mwh"].iloc[0]
    dagitim = sanayi.loc[sanayi["baglanti"] == "dagitim", "tuketim_mwh"].iloc[0]
    assert iletim == pytest.approx(150000.0)
    assert dagitim == pytest.approx(90000.0)
    assert df["il_kodu"].iloc[0] == 26  # Eskişehir plakası


def test_kaynak_matrisi_ayni_kanonige_eslenen_sutunlari_toplar(
    wb: openpyxl.Workbook,
) -> None:
    """Akarsu+Barajlı → Hidrolik TEK satırda toplanmalı (gerçek dosyada bulunan hata)."""
    df = parser.tablo1_kurulu_guc_oku(wb["Tablo 1"], 202601, "Tablo 1")
    hidrolik = df[df["kaynak"] == "Hidrolik"]
    assert len(hidrolik) == 1  # ayrı satır DEĞİL
    assert hidrolik["kurulu_guc_mw"].iloc[0] == pytest.approx(200.0)  # 80+120
    assert bool(hidrolik["yenilenebilir"].iloc[0]) is True
    assert (df["il_kodu"] == 26).all()


def test_lisanssiz_kurulu_guc(wb: openpyxl.Workbook) -> None:
    df = parser.tablo4_lisanssiz_kurulu_guc_oku(wb["Tablo 4"], 202601, "Tablo 4")
    assert (df["lisans"] == "Lisanssız").all()
    assert len(df) == 2


def test_tablo2_kaynak_toplam_doğalgaz_yazim_varyanti(wb: openpyxl.Workbook) -> None:
    """T2'de 'Doğalgaz' (boşluksuz) yazımı da 'Doğal Gaz' kanonik ismine eşlenmeli."""
    df = parser.tablo2_uretim_kaynak_oku(wb["Tablo 2"], 202601)
    assert "il" not in df.columns  # T2 il boyutu içermez
    dogal_gaz = df[df["kaynak"] == "Doğal Gaz"]
    assert len(dogal_gaz) == 1
    assert dogal_gaz["uretim_mwh"].iloc[0] == pytest.approx(400000.0)
    assert "TOPLAM" not in {
        k.upper() for k in df["kaynak"]
    }  # toplam satırı dahil edilmemeli
    assert len(df) == 5


def test_tablo3_il_toplam(wb: openpyxl.Workbook) -> None:
    df = parser.tablo3_uretim_il_oku(wb["Tablo 3"], 202601)
    assert len(df) == 1
    assert df["il_kodu"].iloc[0] == 26
    assert df["uretim_mwh"].iloc[0] == pytest.approx(880000.0)


def test_tablo5_6_lisanssiz_uretim(wb: openpyxl.Workbook) -> None:
    kaynak = parser.tablo5_lisanssiz_uretim_kaynak_oku(wb["Tablo 5"], 202601)
    assert len(kaynak) == 2
    il = parser.tablo6_lisanssiz_uretim_il_oku(wb["Tablo 6"], 202601)
    assert il["uretim_mwh"].iloc[0] == pytest.approx(6000.0)


def test_tablo7_ulke_geneli_mutabakat(wb: openpyxl.Workbook) -> None:
    df = parser.tablo7_faturalanan_tur_oku(wb["Tablo 7"], 202601)
    assert len(df) == 5
    assert (df["il"] == "TÜRKİYE").all()
    assert df["il_kodu"].isna().all()  # ülke geneli, gerçek il değil
    assert df["tuketim_mwh"].sum() == pytest.approx(445000.0)


def test_tablo9_ulke_geneli_mutabakat(wb: openpyxl.Workbook) -> None:
    df = parser.tablo9_abone_tur_oku(wb["Tablo 9"], 202601)
    assert len(df) == 5
    assert df["abone_sayisi"].sum() == pytest.approx(256300.0)


def test_tablo10_il_bazinda_ve_istanbul_birlestirme(wb: openpyxl.Workbook) -> None:
    df = parser.tablo10_abone_il_oku(wb["Tablo 10"], 202601)
    eskisehir = df[df["il_kodu"] == 26]
    assert len(eskisehir) == 5
    assert eskisehir["abone_sayisi"].sum() == pytest.approx(256300.0)

    # İstanbul (Anadolu) + İstanbul (Avrupa) -> tek "İstanbul" satırına toplanmalı
    istanbul = df[df["il_kodu"] == 34]
    assert len(istanbul) == 1
    assert istanbul["il"].iloc[0] == "İstanbul"
    assert istanbul["abone_sayisi"].iloc[0] == pytest.approx(150.0)  # 100+50


def test_ucdan_uca_golden_kpi_ile_esler(wb: openpyxl.Workbook) -> None:
    """Parser çıktısı → kpi.py doğrulama+hesap → golden/expected/kpi_expected.json."""
    kurulu = parser.tablo1_kurulu_guc_oku(wb["Tablo 1"], 202601, "Tablo 1")
    uretim_kaynak = parser.tablo2_uretim_kaynak_oku(wb["Tablo 2"], 202601)
    uretim = kpi.dogrula_uretim(
        parser.uretim_kaynak_birlestir(kurulu, uretim_kaynak)
    ).kabul
    abone = kpi.dogrula_abone(parser.tablo10_abone_il_oku(wb["Tablo 10"], 202601)).kabul
    # Yalnızca Eskişehir'i al (İstanbul test satırlarını golden karşılaştırmadan hariç tut)
    abone = abone[abone["il_kodu"] == 26].reset_index(drop=True)
    tuketim = kpi.dogrula_tuketim(
        parser.tablo11_tuketim_oku(wb["Tablo 11"], 202601, "Tablo 11")
    ).kabul

    with GOLDEN_BEKLENEN.open(encoding="utf-8") as f:
        beklenen = json.load(f)

    assert kpi.kpi_01_kurulu_guc(uretim) == pytest.approx(beklenen["KPI-01"])
    assert kpi.kpi_02_toplam_uretim(uretim) == pytest.approx(beklenen["KPI-02"])
    assert kpi.kpi_03_yenilenebilir_pay(uretim) == pytest.approx(
        beklenen["KPI-03"], rel=0.005
    )
    assert kpi.kpi_06_hhi(uretim) == pytest.approx(beklenen["KPI-06"], rel=0.005)
    assert kpi.kpi_08_toplam_tuketim(tuketim) == pytest.approx(beklenen["KPI-08"])
    assert kpi.kpi_09_grup_payi(tuketim, "Mesken") == pytest.approx(
        beklenen["KPI-09_mesken"], rel=0.005
    )
    assert kpi.kpi_10_abone_basi(tuketim, abone, "Mesken") == pytest.approx(
        beklenen["KPI-10_mesken"]
    )
    p0_2 = kpi.p0_2_sanayi(tuketim)
    for alan, deger in beklenen["P0-2"].items():
        assert p0_2[alan] == pytest.approx(deger, rel=0.005)


@pytest.mark.parametrize(
    ("girdi", "beklenen"),
    [
        ("1.432,404", 1432.404),
        ("0,5", 0.5),
        ("", None),
        ("-", None),
        (None, None),
        (1200, 1200.0),
        (1200.5, 1200.5),
    ],
)
def test_parse_sayi(girdi: object, beklenen: float | None) -> None:
    sonuc = parser.parse_sayi(girdi)
    if beklenen is None:
        assert sonuc is None
    else:
        assert sonuc == pytest.approx(beklenen)


def test_normalize_label_turkce_sadelestirme() -> None:
    assert parser.normalize_label("Sanayi-İLETİM") == "SANAYI-ILETIM"
    assert parser.normalize_label("  Eskişehir  ") == "ESKISEHIR"
    assert parser.normalize_label(None) == ""


def test_il_kodu_bul() -> None:
    assert parser.il_kodu_bul("Eskişehir") == 26
    assert parser.il_kodu_bul("eskisehir") == 26
    assert parser.il_kodu_bul("Bilinmeyen İl XYZ") is None
    assert parser.il_kodu_bul("İstanbul (Anadolu)") == 34
    assert parser.il_kodu_bul("İstanbul (Avrupa)") == 34


def test_kaynak_esle() -> None:
    assert parser.kaynak_esle("Akarsu") == ("Hidrolik", True)
    assert parser.kaynak_esle("Barajlı") == ("Hidrolik", True)
    assert parser.kaynak_esle("İthal Kömür") == ("İthal Kömür", False)
    assert parser.kaynak_esle("Asfaltit Kömür") == ("Asfaltit", False)
    assert parser.kaynak_esle("Doğalgaz") == ("Doğal Gaz", False)  # boşluksuz varyant
    assert parser.kaynak_esle("Taş Kömür") == ("Taş Kömürü", False)
    assert parser.kaynak_esle("Motorin") == ("Motorin", False)
    assert parser.kaynak_esle("Nafta") == ("Nafta", False)
    assert parser.kaynak_esle("Uzay Enerjisi") is None


def test_grup_esle() -> None:
    assert (
        parser.grup_esle("Kamu ve Özel Hizmetler Sektörü ile Diğer")
        == "Kamu ve Özel Hizmetler"
    )
    assert parser.grup_esle("Tarımsal Faaliyetler") == "Tarımsal"
    assert parser.grup_esle("Bilinmeyen Grup") is None


def test_eksik_tablolari_bul_birlesik_sayfalar() -> None:
    """Gerçek dosyada birden fazla tablo aynı sayfada olabilir (ör. 'Tablo 2-3')."""
    mevcut_sayfalar = [
        "Tablo 1",
        "Tablo 2-3",
        "Tablo 4",
        "Tablo 5-6",
        "Tablo 7-8",
        "Tablo 9-10",
        "Tablo 11",
        "Tablo 12",
        "Tablo 13",
    ]
    gerekli = [f"Tablo {n}" for n in range(1, 14)]
    assert parser.eksik_tablolari_bul(mevcut_sayfalar, gerekli) == []

    eksik = parser.eksik_tablolari_bul(mevcut_sayfalar[:-1], gerekli)
    assert eksik == ["Tablo 13"]


def test_mutabakat_kontrol() -> None:
    assert parser.mutabakat_kontrol(1000.0, 1004.0) is True  # %0.4 sapma
    assert parser.mutabakat_kontrol(1000.0, 1010.0) is False  # %1.0 sapma
    assert parser.mutabakat_kontrol(0.0, 0.0) is True


def test_uretim_kaynak_birlestir_ulusal_grain() -> None:
    """İl×kaynak (T1) kaynak bazında toplanır, ülke geneli üretimle (T2) birleşir."""
    kurulu = pd.DataFrame(
        [
            {
                "il": "A",
                "il_kodu": 1,
                "tarih_id": 202601,
                "kaynak": "Rüzgar",
                "yenilenebilir": True,
                "lisans": "Lisanslı",
                "kurulu_guc_mw": 10.0,
            },
            {
                "il": "B",
                "il_kodu": 2,
                "tarih_id": 202601,
                "kaynak": "Rüzgar",
                "yenilenebilir": True,
                "lisans": "Lisanslı",
                "kurulu_guc_mw": 20.0,
            },
        ]
    )
    uretim = pd.DataFrame(
        [
            {
                "kaynak": "Rüzgar",
                "yenilenebilir": True,
                "lisans": "Lisanslı",
                "tarih_id": 202601,
                "uretim_mwh": 5000.0,
            }
        ]
    )
    birlesik = parser.uretim_kaynak_birlestir(kurulu, uretim)
    assert len(birlesik) == 1
    assert birlesik["kurulu_guc_mw"].iloc[0] == pytest.approx(
        30.0
    )  # 10+20, il boyutu kayboldu
    assert birlesik["uretim_mwh"].iloc[0] == pytest.approx(5000.0)


def test_tur_esle() -> None:
    assert parser.tur_esle("Serbest Tüketici") == "Serbest Tuketici"
    assert parser.tur_esle("ST Olma Hakkı Bulunmayan Aboneler") == (
        "ST Olma Hakki Bulunmayan Aboneler"
    )
    assert parser.tur_esle("ST Olma Hakkını Kullanmayan Aboneler") == (
        "ST Olma Hakkini Kullanmayan Aboneler"
    )
    assert parser.tur_esle("Bilinmeyen Tür") is None


def test_grup_esle_kamu_ozel_kisaltma() -> None:
    """T13'te 'Kamu/Özel ' (sondaki boşluklu) kısaltması kullanılıyor."""
    assert parser.grup_esle("Kamu/Özel ") == "Kamu ve Özel Hizmetler"
    assert parser.grup_esle("Kamu/Özel") == "Kamu ve Özel Hizmetler"


def test_tablo13_serbest_tuketici(wb: openpyxl.Workbook) -> None:
    df = parser.tablo13_serbest_tuketici_oku(wb["Tablo 13"], 202601)

    # "İl Toplam" satırları hiç üretilmemeli: 3 tur × 5 grup (Eskişehir) +
    # 1 tur × 5 grup (Adana) + 1 tur × 5 grup (İstanbul, Anadolu+Avrupa
    # toplanmış) = 25 satır.
    assert len(df) == 25
    assert set(df["tur"].unique()) == {
        "Serbest Tuketici",
        "ST Olma Hakki Bulunmayan Aboneler",
        "ST Olma Hakkini Kullanmayan Aboneler",
    }
    assert set(df["grup"].unique()) == {
        "Mesken",
        "Sanayi",
        "Kamu ve Özel Hizmetler",
        "Tarımsal",
        "Aydınlatma",
    }

    eskisehir = df[(df["il_kodu"] == 26) & (df["tur"] == "Serbest Tuketici")]
    mesken = eskisehir[eskisehir["grup"] == "Mesken"].iloc[0]
    assert mesken["tuketim_mwh"] == pytest.approx(100.0)
    assert mesken["tuketici_sayisi"] == pytest.approx(5.0)

    # Boş hücre = NULL (0 DEĞİL)
    tarimsal = eskisehir[eskisehir["grup"] == "Tarımsal"].iloc[0]
    assert pd.isna(tarimsal["tuketim_mwh"])
    assert pd.isna(tarimsal["tuketici_sayisi"])

    # Negatif değer parser seviyesinde reddedilmez (bu dogrula_serbest_tuketici'nin işi);
    # ham değer olduğu gibi geçmeli.
    kullanmayan = df[
        (df["il_kodu"] == 26)
        & (df["tur"] == "ST Olma Hakkini Kullanmayan Aboneler")
        & (df["grup"] == "Mesken")
    ].iloc[0]
    assert kullanmayan["tuketim_mwh"] == pytest.approx(-5.0)

    # İleri doldurma: Adana'nın tek satırı da doğru il'e bağlanmalı.
    # "il" metni artık kanonik ada normalize ediliyor (T9/T10 ile aynı desen,
    # bkz. tablo13_serbest_tuketici_oku'daki İstanbul Anadolu/Avrupa toplama
    # düzeltmesi) - ham hücre metni ("ADANA") değil il_adi_kanonik() çıktısı.
    adana = df[df["il_kodu"] == 1]
    assert len(adana) == 5
    assert (adana["il"] == "Adana").all()

    # İstanbul (Anadolu) + İstanbul (Avrupa) -> tek "İstanbul" satırına
    # toplanmalı, hiçbiri sessizce kaybolmamalı.
    istanbul = df[df["il_kodu"] == 34]
    assert len(istanbul) == 5
    assert (istanbul["il"] == "İstanbul").all()
    mesken_ist = istanbul[istanbul["grup"] == "Mesken"].iloc[0]
    assert mesken_ist["tuketim_mwh"] == pytest.approx(110.0)  # 10 + 100
    assert mesken_ist["tuketici_sayisi"] == pytest.approx(11.0)  # 1 + 10
    aydinlatma_ist = istanbul[istanbul["grup"] == "Aydınlatma"].iloc[0]
    assert aydinlatma_ist["tuketim_mwh"] == pytest.approx(550.0)  # 50 + 500
    assert aydinlatma_ist["tuketici_sayisi"] == pytest.approx(55.0)  # 5 + 50


def test_dogrula_serbest_tuketici_red_ve_karantina() -> None:
    df = pd.DataFrame(
        [
            {
                "il_kodu": 26,
                "tarih_id": 202601,
                "tur": "Serbest Tuketici",
                "grup": "Mesken",
                "tuketim_mwh": 100.0,
                "tuketici_sayisi": 5.0,
            },
            {
                "il_kodu": 26,
                "tarih_id": 202601,
                "tur": "Serbest Tuketici",
                "grup": "Mesken",
                "tuketim_mwh": -5.0,
                "tuketici_sayisi": 1.0,
            },
            {
                "il_kodu": 26,
                "tarih_id": 202601,
                "tur": "Bilinmeyen Tur",
                "grup": "Mesken",
                "tuketim_mwh": 10.0,
                "tuketici_sayisi": 1.0,
            },
            {
                "il_kodu": 26,
                "tarih_id": 202601,
                "tur": "Serbest Tuketici",
                "grup": "Bilinmeyen Grup",
                "tuketim_mwh": 10.0,
                "tuketici_sayisi": 1.0,
            },
        ]
    )
    sonuc = kpi.dogrula_serbest_tuketici(df)
    assert len(sonuc.kabul) == 1
    assert len(sonuc.red) == 1
    assert len(sonuc.karantina) == 2
