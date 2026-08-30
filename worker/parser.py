"""EPP — EPDK aylık ek (xlsx) anchor-tabanlı parser (Faz 0).

Kaynak: dokumanlar/05_kaynak_dosya_sozlesmesi.md (Ek F). Parser sabit hücreye
güvenmez; değişmez etiketleri (tablo başlığı, sütun adı, 'TÜRKİYE' satırı)
arar.

Bu sürüm 2026 Ocak EPDK Elektrik Piyasası Sektör Raporu Ek'i (gerçek dosya)
ile doğrulanmıştır. Doğrulama gerçek dosyayı, dokümandaki varsayımlardan
farklı olan şu noktaları ortaya çıkardı:

- T1/T4 (kurulu güç): aynı kanonik kaynağa eşlenen BİRDEN FAZLA ham sütun var
  (Akarsu+Barajlı→Hidrolik, Doğal Gaz+LNG→Doğal Gaz) — toplanmalı, ayrı satır
  ÜRETİLMEMELİ.
- Kaynak/il başlık etiketleri tablolar arası tutarsız yazılmış (ör. "İLLER"
  vs "İL" vs "İl Adı"; "Doğal Gaz" vs "Doğalgaz"; "Taş Kömür" vs "Taş
  Kömürü"; "Asfaltit Kömür" vs "ASFALTİT").
- T2/T3, T5/T6, T7/T8, T9/T10 dokümanın ima ettiği gibi il×kategori matrisi
  DEĞİL: her çift, biri ülke geneli (kaynak/tür bazında, il YOK — T2/T5/T7/T9)
  diğeri il bazında toplam (kategori kırılımı YOK — T3/T6/T8/T10) olan iki
  AYRI, tek-boyutlu tablodur. İl × kaynak/kategori kesişimi (fact_uretim'in
  beklediği grain) aylık raporda mevcut değildir — yalnız kurulu_guc_mw bu
  detayda var (T1/T4), uretim_mwh değil.
- T8, T11 ile birebir aynı il×tüketici-grubu verisini tekrarlıyor (T11 ayrıca
  Sanayi'yi iletim/dağıtım olarak ayırıyor) → T8 ayrıca implemente edilmedi,
  T11 tek başına yeterli. T7/T9 yalnız ülke geneli mutabakat satırları
  (fact tablosuna yazılmaz, yalnız mutabakat_kontrol için).
- T12'nin gerçek boyutu "il" değil "dağıtım şirketi" (21 şirket + ulusal
  'İletimden Bağlı Tüketiciler' satırı, yalnız Sanayi) ve T11 ile birebir
  redundant (Genel Toplam + Sanayi-İletim rakamları eşleşiyor) — bu yüzden
  parse edilmiyor (bkz. dokumanlar/05_kaynak_dosya_sozlesmesi.md T12 notu).
- T13 (serbest tüketici) gerçek dosyayla doğrulandı (2026-08-30): grain
  dokümanın ima ettiğinden farklı, her (il, tur) AYRICA 5 tüketici grubuna
  bölünmüş (bkz. migration 20260819_0006, tablo13_serbest_tuketici_oku).
  Gerçek "tur" değerleri 'Lisanslı'/'Lisanssız' DEĞİL: 'Serbest Tüketici',
  'ST Olma Hakkı Bulunmayan Aboneler', 'ST Olma Hakkını Kullanmayan
  Aboneler'. Yerleşim iki paralel bloktur (Tüketim Miktarı + Tüketici
  Sayısı, aynı 5 grup adı iki kez).
"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

import pandas as pd

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Normalizasyon (dokumanlar/05_kaynak_dosya_sozlesmesi.md — Çapa Tabanlı Okuma)
# ---------------------------------------------------------------------------

_TR_SADE = str.maketrans(
    {
        "ç": "C",
        "Ç": "C",
        "ğ": "G",
        "Ğ": "G",
        "ı": "I",
        "İ": "I",
        "i": "I",
        "ö": "O",
        "Ö": "O",
        "ş": "S",
        "Ş": "S",
        "ü": "U",
        "Ü": "U",
    }
)


def normalize_label(deger: object) -> str:
    """Trim + BÜYÜK harf + Türkçe sadeleştir (İ→I). Eşleme/çapa arama içindir."""
    if deger is None:
        return ""
    metin = str(deger).translate(_TR_SADE).upper()
    return " ".join(metin.split())


def _sade_anahtar(deger: object) -> str:
    """normalize_label + TÜM boşlukları at. 'Doğal Gaz' ile 'Doğalgaz' aynı anahtara düşer."""
    return normalize_label(deger).replace(" ", "")


def parse_sayi(deger: object) -> float | None:
    """'1.432,404' → 1432.404. Boş hücre → None (0 DEĞİL)."""
    if deger is None:
        return None
    if isinstance(deger, int | float):
        return float(deger)
    metin = str(deger).strip()
    if metin in ("", "-", "—", "n/a", "N/A"):
        return None
    metin = metin.replace(".", "").replace(",", ".")
    try:
        return float(metin)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# İl plaka kodu (81 il — resmi plaka numaralandırması)
# ---------------------------------------------------------------------------

_IL_PLAKA_HAM: dict[str, int] = {
    "Adana": 1,
    "Adıyaman": 2,
    "Afyonkarahisar": 3,
    "Ağrı": 4,
    "Amasya": 5,
    "Ankara": 6,
    "Antalya": 7,
    "Artvin": 8,
    "Aydın": 9,
    "Balıkesir": 10,
    "Bilecik": 11,
    "Bingöl": 12,
    "Bitlis": 13,
    "Bolu": 14,
    "Burdur": 15,
    "Bursa": 16,
    "Çanakkale": 17,
    "Çankırı": 18,
    "Çorum": 19,
    "Denizli": 20,
    "Diyarbakır": 21,
    "Edirne": 22,
    "Elazığ": 23,
    "Erzincan": 24,
    "Erzurum": 25,
    "Eskişehir": 26,
    "Gaziantep": 27,
    "Giresun": 28,
    "Gümüşhane": 29,
    "Hakkari": 30,
    "Hatay": 31,
    "Isparta": 32,
    "Mersin": 33,
    "İstanbul": 34,
    "İzmir": 35,
    "Kars": 36,
    "Kastamonu": 37,
    "Kayseri": 38,
    "Kırklareli": 39,
    "Kırşehir": 40,
    "Kocaeli": 41,
    "Konya": 42,
    "Kütahya": 43,
    "Malatya": 44,
    "Manisa": 45,
    "Kahramanmaraş": 46,
    "Mardin": 47,
    "Muğla": 48,
    "Muş": 49,
    "Nevşehir": 50,
    "Niğde": 51,
    "Ordu": 52,
    "Rize": 53,
    "Sakarya": 54,
    "Samsun": 55,
    "Siirt": 56,
    "Sinop": 57,
    "Sivas": 58,
    "Tekirdağ": 59,
    "Tokat": 60,
    "Trabzon": 61,
    "Tunceli": 62,
    "Şanlıurfa": 63,
    "Uşak": 64,
    "Van": 65,
    "Yozgat": 66,
    "Zonguldak": 67,
    "Aksaray": 68,
    "Bayburt": 69,
    "Karaman": 70,
    "Kırıkkale": 71,
    "Batman": 72,
    "Şırnak": 73,
    "Bartın": 74,
    "Ardahan": 75,
    "Iğdır": 76,
    "Yalova": 77,
    "Karabük": 78,
    "Kilis": 79,
    "Osmaniye": 80,
    "Düzce": 81,
}

IL_PLAKA: dict[str, int] = {
    normalize_label(ad): kod for ad, kod in _IL_PLAKA_HAM.items()
}
_IL_ADI_KANONIK: dict[int, str] = {kod: ad for ad, kod in _IL_PLAKA_HAM.items()}


def il_adi_kanonik(il_kodu: int | None) -> str | None:
    if il_kodu is None:
        return None
    return _IL_ADI_KANONIK.get(int(il_kodu))


# T9/T10'da tüketici sayısı İstanbul için iki dağıtım bölgesine (Anadolu/Avrupa)
# bölünmüş ayrı satırlar olarak geliyor; ikisi de plaka 34'e karşılık gelir.
# (_uzun_format_grup_oku bu ikisini il_kodu bazında toplayarak tek satıra indirir.)
IL_PLAKA[normalize_label("İstanbul (Anadolu)")] = 34
IL_PLAKA[normalize_label("İstanbul (Avrupa)")] = 34


def il_kodu_bul(il_adi: object) -> int | None:
    return IL_PLAKA.get(normalize_label(il_adi))


# ---------------------------------------------------------------------------
# Kaynak Türü Eşleme (dokumanlar/05_kaynak_dosya_sozlesmesi.md + gerçek dosya)
# ---------------------------------------------------------------------------
# Anahtar _sade_anahtar (boşluksuz) — "Doğal Gaz"/"Doğalgaz" gibi yazım
# varyasyonlarını otomatik yakalamak için. Yine de görülen tüm varyantlar
# açıkça listelenir (sessiz veri kaybını önlemek amacıyla, bkz. modül notu).

KAYNAK_ESLEME: dict[str, tuple[str, bool]] = {
    _sade_anahtar(etiket): (kanonik, yenilenebilir)
    for etiketler, kanonik, yenilenebilir in [
        (["Akarsu", "Barajlı", "Hidrolik"], "Hidrolik", True),
        (["Rüzgar"], "Rüzgar", True),
        (["Güneş", "Güneş (Fotovoltaik)"], "Güneş", True),
        (["Jeotermal"], "Jeotermal", True),
        (["Biyokütle"], "Biyokütle", True),
        (["Doğal Gaz", "Doğalgaz", "LNG"], "Doğal Gaz", False),
        (["İthal Kömür"], "İthal Kömür", False),
        (["Linyit"], "Linyit", False),
        (["Taş Kömür", "Taş Kömürü"], "Taş Kömürü", False),
        (["Asfaltit", "Asfaltit Kömür"], "Asfaltit", False),
        (["Fuel Oil"], "Fuel Oil", False),
        (["Motorin"], "Motorin", False),
        (["Nafta"], "Nafta", False),
    ]
    for etiket in etiketler
}


def kaynak_esle(etiket: object) -> tuple[str, bool] | None:
    return KAYNAK_ESLEME.get(_sade_anahtar(etiket))


# ---------------------------------------------------------------------------
# Tüketici Grubu Eşleme (uzun-format tablolarda hücre değeri olarak gelir)
# ---------------------------------------------------------------------------

GRUP_ESLEME: dict[str, str] = {
    _sade_anahtar(etiket): kanonik
    for etiketler, kanonik in [
        (["Aydınlatma"], "Aydınlatma"),
        (
            [
                "Kamu ve Özel Hizmetler",
                "Kamu ve Özel Hizmetler Sektörü ile Diğer",
                "Kamu/Özel",  # T13'te kısaltılmış yazım (_sade_anahtar zaten trim eder)
            ],
            "Kamu ve Özel Hizmetler",
        ),
        (["Mesken"], "Mesken"),
        (["Sanayi"], "Sanayi"),
        (["Tarımsal", "Tarımsal Faaliyetler"], "Tarımsal"),
    ]
    for etiket in etiketler
}


def grup_esle(etiket: object) -> str | None:
    return GRUP_ESLEME.get(_sade_anahtar(etiket))


# ---------------------------------------------------------------------------
# Serbest Tüketici Türü Eşleme (T13) — gerçek Türkçe etiket → migration
# 20260819_0006'daki (Türkçe karaktersiz) kanonik fact_serbest_tuketici.tur
# CHECK değerleri. 'Lisanslı'/'Lisanssız' DEĞİL — dokumanlar/05 2026-08-30 notu.
# ---------------------------------------------------------------------------

TUR_ESLEME: dict[str, str] = {
    _sade_anahtar(etiket): kanonik
    for etiket, kanonik in [
        ("Serbest Tüketici", "Serbest Tuketici"),
        ("ST Olma Hakkı Bulunmayan Aboneler", "ST Olma Hakki Bulunmayan Aboneler"),
        (
            "ST Olma Hakkını Kullanmayan Aboneler",
            "ST Olma Hakkini Kullanmayan Aboneler",
        ),
    ]
}


def tur_esle(etiket: object) -> str | None:
    return TUR_ESLEME.get(_sade_anahtar(etiket))


# ---------------------------------------------------------------------------
# Çapa (Anchor) arama
# ---------------------------------------------------------------------------


def bul_capa(ws: Worksheet, etiket: str, min_row: int = 1) -> tuple[int, int] | None:
    """Normalize edilmiş hücre değeri `etiket`i içeren ilk hücrenin (satır, sütun) konumu."""
    hedef = normalize_label(etiket)
    if not hedef:
        return None
    for row in ws.iter_rows(min_row=min_row):
        for cell in row:
            if hedef in normalize_label(cell.value):
                return (cell.row, cell.column)
    return None


def _satirda_kolon_bul(
    ws: Worksheet, satir: int, etiket: str, min_col: int = 1, max_col: int = 60
) -> int | None:
    hedef = normalize_label(etiket)
    if not hedef:
        return None
    for col in range(min_col, max_col + 1):
        if hedef in normalize_label(ws.cell(row=satir, column=col).value):
            return col
    return None


_DURDURMA_ETIKETLERI = {"TURKIYE", "GENEL TOPLAM", "TOPLAM"}

# Gerçek dosyada il-kolonu başlığı tablo tablo farklı yazılmış (İLLER/İL/İl Adı).
# Substring aramasında "İLLER" ile arasak "İL" içeren başlıkları KAÇIRIRIZ (hedef
# aranan metinden uzun olamaz) — bu yüzden tam eşleşme ile birkaç varyant deneriz.
_IL_BASLIK_VARYANTLARI = ("İLLER", "İL", "İl Adı")


def _il_matrisi_oku(ws: Worksheet, tablo_etiketi: str) -> tuple[int, int] | None:
    """Tablo çapasını ve altındaki il başlık satırını bulur → (baslik_satir, il_sutun)."""
    capa = bul_capa(ws, tablo_etiketi)
    if capa is None:
        return None
    tablo_satir, _ = capa
    for satir in range(tablo_satir, tablo_satir + 10):
        for varyant in _IL_BASLIK_VARYANTLARI:
            hedef = normalize_label(varyant)
            for col in range(1, 60):
                if normalize_label(ws.cell(row=satir, column=col).value) == hedef:
                    return (satir, col)
    return None


def _veri_satirlarini_gez(ws: Worksheet, baslik_satir: int, il_sutun: int):
    """(satir_no, il_adi) ikililerini 'TÜRKİYE/TOPLAM' veya boş hücreye kadar üretir."""
    satir = baslik_satir + 1
    while True:
        il_deger = ws.cell(row=satir, column=il_sutun).value
        if il_deger is None or str(il_deger).strip() == "":
            return
        if normalize_label(il_deger) in _DURDURMA_ETIKETLERI:
            return
        yield satir, str(il_deger).strip()
        satir += 1


# ---------------------------------------------------------------------------
# T11 — Tüketim (İLETİM/DAĞITIM) — P0-2 KRİTİK (il × grup matrisi)
# ---------------------------------------------------------------------------

_T11_GRUP_KOLONLARI = [
    ("Aydınlatma", "Aydınlatma", "dagitim"),
    ("Kamu", "Kamu ve Özel Hizmetler", "dagitim"),
    ("Mesken", "Mesken", "dagitim"),
    ("Sanayi-DAGITIM", "Sanayi", "dagitim"),
    ("Sanayi-ILETIM", "Sanayi", "iletim"),
    ("Tarımsal", "Tarımsal", "dagitim"),
]


def tablo11_tuketim_oku(
    ws: Worksheet, tarih_id: int, tablo_etiketi: str = "Tablo 11"
) -> pd.DataFrame:
    """T11: il × (Aydınlatma, Kamu, Mesken, Sanayi-DAĞITIM, Sanayi-İLETİM, Tarımsal).

    P0-2: Sanayi-DAĞITIM ve Sanayi-İLETİM AYRI satır (grup='Sanayi', baglanti farklı).
    Tek başına yeterli (T8 ile aynı il×grup verisini, ayrıca sanayi ayrımıyla verir).
    """
    konum = _il_matrisi_oku(ws, tablo_etiketi)
    if konum is None:
        return pd.DataFrame(
            columns=["il", "il_kodu", "tarih_id", "grup", "baglanti", "tuketim_mwh"]
        )
    baslik_satir, il_sutun = konum

    kolon_indeksleri = [
        (_satirda_kolon_bul(ws, baslik_satir, arama_etiketi), grup, baglanti)
        for arama_etiketi, grup, baglanti in _T11_GRUP_KOLONLARI
    ]

    satirlar = []
    for satir_no, il_adi in _veri_satirlarini_gez(ws, baslik_satir, il_sutun):
        for kolon, grup, baglanti in kolon_indeksleri:
            if kolon is None:
                continue
            deger = parse_sayi(ws.cell(row=satir_no, column=kolon).value)
            satirlar.append(
                {
                    "il": il_adi,
                    "il_kodu": il_kodu_bul(il_adi),
                    "tarih_id": tarih_id,
                    "grup": grup,
                    "baglanti": baglanti,
                    "tuketim_mwh": deger,
                }
            )
    return pd.DataFrame(
        satirlar,
        columns=["il", "il_kodu", "tarih_id", "grup", "baglanti", "tuketim_mwh"],
    )


# ---------------------------------------------------------------------------
# T7/T9 (ülke geneli, mutabakat) + T8/T10 (il bazında) — UZUN format
# ---------------------------------------------------------------------------
# Gerçek yerleşim: (il_adi_veya_bos, Tüketici Grubu, Miktar/Sayı) üç sütun.
# İl adı yalnız o ilin ilk satırında yazılı (birleştirilmiş hücre stili);
# sonraki 4 satırda boş → bir önceki il'e ait kabul edilir (ileri doldurma).
# Ülke geneli tablolarda (T7/T9) tek 'TÜRKİYE' grubu vardır. Her il/TÜRKİYE
# grubunun sonunda bir alt-toplam satırı var ('Toplam' ya da 'İl Toplam') —
# bu satır veri değil, atlanır.

_UZUN_FORMAT_ATLA = {"TOPLAM", "IL TOPLAM", "GENEL TOPLAM"}


def _uzun_format_grup_oku(
    ws: Worksheet,
    tablo_etiketi: str,
    tarih_id: int,
    deger_kolon_adi: str,
    baglanti: str | None = None,
) -> pd.DataFrame:
    kolonlar = ["il", "il_kodu", "tarih_id", "grup", deger_kolon_adi]
    if baglanti is not None:
        kolonlar.insert(4, "baglanti")

    capa = bul_capa(ws, tablo_etiketi)
    if capa is None:
        return pd.DataFrame(columns=kolonlar)
    tablo_satir, _ = capa

    baslik_satir = None
    grup_kolon = None
    for satir in range(tablo_satir, tablo_satir + 8):
        kolon = _satirda_kolon_bul(ws, satir, "Tüketici Grubu")
        if kolon is not None:
            baslik_satir, grup_kolon = satir, kolon
            break
    if baslik_satir is None or grup_kolon is None:
        return pd.DataFrame(columns=kolonlar)
    il_kolon = grup_kolon - 1
    deger_kolon = grup_kolon + 1

    satirlar = []
    mevcut_il: str | None = None
    satir = baslik_satir + 1
    while True:
        grup_ham = ws.cell(row=satir, column=grup_kolon).value
        if grup_ham is None or str(grup_ham).strip() == "":
            break
        il_ham = ws.cell(row=satir, column=il_kolon).value
        if il_ham is not None and str(il_ham).strip() != "":
            mevcut_il = str(il_ham).strip()
        if normalize_label(grup_ham) not in _UZUN_FORMAT_ATLA:
            grup = grup_esle(grup_ham)
            if grup is not None:
                deger = parse_sayi(ws.cell(row=satir, column=deger_kolon).value)
                satir_dict = {
                    "il": mevcut_il,
                    "il_kodu": il_kodu_bul(mevcut_il),
                    "tarih_id": tarih_id,
                    "grup": grup,
                    deger_kolon_adi: deger,
                }
                if baglanti is not None:
                    satir_dict["baglanti"] = baglanti
                satirlar.append(satir_dict)
        satir += 1

    df = pd.DataFrame(satirlar, columns=kolonlar)
    if df.empty:
        return df
    # İstanbul gibi birden fazla dağıtım bölgesine bölünmüş il'ler aynı il_kodu'na
    # düşer; "il" metni farklı olduğundan gruplamadan HARİÇ tutulur, tekrar eden
    # (il_kodu, grup) satırları toplanır, sonra kanonik il adı geri eklenir.
    grup_kolonlari = [k for k in kolonlar if k not in (deger_kolon_adi, "il")]
    sonuc = df.groupby(grup_kolonlari, as_index=False, dropna=False)[
        deger_kolon_adi
    ].sum(min_count=1)
    sonuc["il"] = sonuc["il_kodu"].map(
        lambda k: il_adi_kanonik(k) if pd.notna(k) else "TÜRKİYE"
    )
    return sonuc[kolonlar]


def tablo7_faturalanan_tur_oku(ws: Worksheet, tarih_id: int) -> pd.DataFrame:
    """T7: ülke geneli, tüketici grubu bazında tüketim (MUTABAKAT içindir; il yok)."""
    return _uzun_format_grup_oku(ws, "Tablo 7", tarih_id, "tuketim_mwh")


def tablo9_abone_tur_oku(ws: Worksheet, tarih_id: int) -> pd.DataFrame:
    """T9: ülke geneli, tüketici grubu bazında abone sayısı (MUTABAKAT içindir; il yok)."""
    return _uzun_format_grup_oku(ws, "Tablo 9", tarih_id, "abone_sayisi")


def tablo10_abone_il_oku(ws: Worksheet, tarih_id: int) -> pd.DataFrame:
    """T10: il × tüketici grubu → abone_sayisi. fact_abone'nin BİRİNCİL kaynağı."""
    return _uzun_format_grup_oku(ws, "Tablo 10", tarih_id, "abone_sayisi")


# ---------------------------------------------------------------------------
# T13 — Serbest tüketici (il × tur × grup, iki paralel değer bloğu)
# ---------------------------------------------------------------------------
# Gerçek yerleşim: 'İl Adı' | 'Elektrik Tüketici Türü' | ['Tüketim Miktarı(MWh)'
# başlığı altında 5 grup sütunu + Toplam] | [aynı 5 grup adı 'Tüketici Sayısı'
# başlığı altında tekrar + Toplam]. İl adı yalnız o ilin ilk satırında yazılı
# (ileri doldurma, T7-T10'daki gibi); her ilin sonunda atlanacak bir
# 'İl Toplam' satırı var. 'Toplam' sütunları grup_esle() ile zaten elenir.

_SERBEST_TUKETICI_ATLA = {"TOPLAM", "IL TOPLAM"}


def _grup_kolonlarini_tara(
    ws: Worksheet, satir: int, baslangic_col: int, bitis_col: int
) -> list[tuple[int, str]]:
    """[baslangic_col, bitis_col] aralığında grup_esle ile eşleşen sütunları toplar."""
    sonuc = []
    for col in range(baslangic_col, bitis_col + 1):
        deger = ws.cell(row=satir, column=col).value
        if deger is None or str(deger).strip() == "":
            continue
        grup = grup_esle(deger)
        if grup is not None:
            sonuc.append((col, grup))
    return sonuc


def tablo13_serbest_tuketici_oku(
    ws: Worksheet, tarih_id: int, tablo_etiketi: str = "Tablo 13"
) -> pd.DataFrame:
    """T13: il (ileri doldurma) × tur × grup → tuketim_mwh + tuketici_sayisi."""
    kolonlar = [
        "il",
        "il_kodu",
        "tarih_id",
        "tur",
        "grup",
        "tuketim_mwh",
        "tuketici_sayisi",
    ]

    capa = bul_capa(ws, tablo_etiketi)
    if capa is None:
        return pd.DataFrame(columns=kolonlar)
    tablo_satir, _ = capa

    baslik_satir = None
    il_kolon = None
    hedef_il_adi = normalize_label("İl Adı")
    for satir in range(tablo_satir, tablo_satir + 6):
        for col in range(1, 5):
            if normalize_label(ws.cell(row=satir, column=col).value) == hedef_il_adi:
                baslik_satir, il_kolon = satir, col
                break
        if baslik_satir is not None:
            break
    if baslik_satir is None or il_kolon is None:
        return pd.DataFrame(columns=kolonlar)

    tur_kolon = il_kolon + 1
    grup_baslik_satir = baslik_satir + 1  # grup adları bir alt satırda tekrar eder

    sayisi_baslangic = _satirda_kolon_bul(
        ws, baslik_satir, "Tüketici Sayısı", min_col=tur_kolon + 1
    )
    if sayisi_baslangic is None:
        return pd.DataFrame(columns=kolonlar)

    tuketim_kolonlari = _grup_kolonlarini_tara(
        ws, grup_baslik_satir, tur_kolon + 1, sayisi_baslangic - 1
    )
    sayisi_kolonlari = {
        grup: col
        for col, grup in _grup_kolonlarini_tara(
            ws, grup_baslik_satir, sayisi_baslangic, sayisi_baslangic + 15
        )
    }

    satirlar = []
    mevcut_il: str | None = None
    satir = grup_baslik_satir + 1
    while True:
        tur_ham = ws.cell(row=satir, column=tur_kolon).value
        if tur_ham is None or str(tur_ham).strip() == "":
            break
        il_ham = ws.cell(row=satir, column=il_kolon).value
        if il_ham is not None and str(il_ham).strip() != "":
            mevcut_il = str(il_ham).strip()
        if normalize_label(tur_ham) not in _SERBEST_TUKETICI_ATLA:
            tur = tur_esle(tur_ham)
            if tur is not None:
                for tuketim_col, grup in tuketim_kolonlari:
                    tuketim_deger = parse_sayi(
                        ws.cell(row=satir, column=tuketim_col).value
                    )
                    sayisi_col = sayisi_kolonlari.get(grup)
                    sayisi_deger = (
                        parse_sayi(ws.cell(row=satir, column=sayisi_col).value)
                        if sayisi_col is not None
                        else None
                    )
                    satirlar.append(
                        {
                            "il": mevcut_il,
                            "il_kodu": il_kodu_bul(mevcut_il),
                            "tarih_id": tarih_id,
                            "tur": tur,
                            "grup": grup,
                            "tuketim_mwh": tuketim_deger,
                            "tuketici_sayisi": sayisi_deger,
                        }
                    )
        satir += 1
    return pd.DataFrame(satirlar, columns=kolonlar)


# ---------------------------------------------------------------------------
# T1/T4 — Kurulu güç (il × kaynak matrisi, lisanslı/lisanssız)
# ---------------------------------------------------------------------------

_URETIM_KOLONLARI = ["il", "il_kodu", "tarih_id", "kaynak", "yenilenebilir", "lisans"]


def _kaynak_matrisi_oku(
    ws: Worksheet,
    tablo_etiketi: str,
    tarih_id: int,
    deger_kolon_adi: str,
    lisans: str = "Lisanslı",
) -> pd.DataFrame:
    """İl × kaynak matrisi (T1/T4). Aynı kanonik kaynağa eşlenen birden fazla ham
    sütun (ör. Akarsu+Barajlı→Hidrolik) TOPLANIR; ayrı satır üretilmez."""
    kolonlar = [*_URETIM_KOLONLARI, deger_kolon_adi]
    konum = _il_matrisi_oku(ws, tablo_etiketi)
    if konum is None:
        return pd.DataFrame(columns=kolonlar)
    baslik_satir, il_sutun = konum

    kaynak_kolonlari = []
    for col in range(il_sutun + 1, il_sutun + 40):
        baslik_deger = ws.cell(row=baslik_satir, column=col).value
        if baslik_deger is None or str(baslik_deger).strip() == "":
            continue
        eslesme = kaynak_esle(baslik_deger)
        if eslesme is not None:
            kaynak_kolonlari.append((col, *eslesme))

    satirlar = []
    for satir_no, il_adi in _veri_satirlarini_gez(ws, baslik_satir, il_sutun):
        toplamlar: dict[
            str, list
        ] = {}  # kaynak -> [toplam, yenilenebilir, veri_var_mi]
        for kolon, kaynak, yenilenebilir in kaynak_kolonlari:
            deger = parse_sayi(ws.cell(row=satir_no, column=kolon).value)
            kayit = toplamlar.setdefault(kaynak, [0.0, yenilenebilir, False])
            if deger is not None:
                kayit[0] += deger
                kayit[2] = True
        for kaynak, (toplam, yenilenebilir, veri_var_mi) in toplamlar.items():
            satirlar.append(
                {
                    "il": il_adi,
                    "il_kodu": il_kodu_bul(il_adi),
                    "tarih_id": tarih_id,
                    "kaynak": kaynak,
                    "yenilenebilir": yenilenebilir,
                    "lisans": lisans,
                    deger_kolon_adi: toplam if veri_var_mi else None,
                }
            )
    return pd.DataFrame(satirlar, columns=kolonlar)


def tablo1_kurulu_guc_oku(
    ws: Worksheet, tarih_id: int, tablo_etiketi: str = "Tablo 1"
) -> pd.DataFrame:
    """T1: il × kaynak → kurulu_guc_mw (lisanslı)."""
    return _kaynak_matrisi_oku(ws, tablo_etiketi, tarih_id, "kurulu_guc_mw", "Lisanslı")


def tablo4_lisanssiz_kurulu_guc_oku(
    ws: Worksheet, tarih_id: int, tablo_etiketi: str = "Tablo 4"
) -> pd.DataFrame:
    """T4: il × kaynak → kurulu_guc_mw (lisanssız)."""
    return _kaynak_matrisi_oku(
        ws, tablo_etiketi, tarih_id, "kurulu_guc_mw", "Lisanssız"
    )


# ---------------------------------------------------------------------------
# T2/T5 (ülke geneli, kaynak bazında) + T3/T6 (il bazında toplam) — üretim
# ---------------------------------------------------------------------------
# Gerçek dosyada üretim (MWh) yalnız İKİ AYRI tek-boyutlu tablo olarak var:
# kaynak bazında ülke toplamı (il yok) VE il bazında toplam (kaynak yok).
# İl×kaynak kesişimi (fact_uretim'in grain'i) hiçbir tabloda mevcut değil.


def tablo_kaynak_toplam_oku(
    ws: Worksheet, tablo_etiketi: str, tarih_id: int, deger_kolon_adi: str, lisans: str
) -> pd.DataFrame:
    """T2/T5: Kaynak Türü × tek ay değeri (ülke toplamı, il YOK)."""
    kolonlar = ["kaynak", "yenilenebilir", "lisans", "tarih_id", deger_kolon_adi]
    capa = bul_capa(ws, tablo_etiketi)
    if capa is None:
        return pd.DataFrame(columns=kolonlar)
    tablo_satir, _ = capa

    baslik_satir = None
    for satir in range(tablo_satir, tablo_satir + 6):
        if _satirda_kolon_bul(ws, satir, "Kaynak Türü") is not None:
            baslik_satir = satir
            break
    if baslik_satir is None:
        return pd.DataFrame(columns=kolonlar)

    satirlar = []
    satir = baslik_satir + 1
    while True:
        etiket = ws.cell(row=satir, column=1).value
        if etiket is None or str(etiket).strip() == "":
            break
        if normalize_label(etiket) in _DURDURMA_ETIKETLERI:
            break
        eslesme = kaynak_esle(etiket)
        if eslesme is not None:
            kaynak, yenilenebilir = eslesme
            deger = parse_sayi(ws.cell(row=satir, column=2).value)
            satirlar.append(
                {
                    "kaynak": kaynak,
                    "yenilenebilir": yenilenebilir,
                    "lisans": lisans,
                    "tarih_id": tarih_id,
                    deger_kolon_adi: deger,
                }
            )
        satir += 1

    df = pd.DataFrame(satirlar, columns=kolonlar)
    if not df.empty:
        # Aynı kanonik kaynağa eşlenen olası tekrar eden satırları topla.
        df = df.groupby(
            ["kaynak", "yenilenebilir", "lisans", "tarih_id"], as_index=False
        )[deger_kolon_adi].sum(min_count=1)
    return df


def tablo_il_toplam_oku(
    ws: Worksheet, tablo_etiketi: str, tarih_id: int, deger_kolon_adi: str, lisans: str
) -> pd.DataFrame:
    """T3/T6: İl × tek ay değeri (toplam üretim, kaynak kırılımı YOK)."""
    kolonlar = ["il", "il_kodu", "tarih_id", "lisans", deger_kolon_adi]
    capa = bul_capa(ws, tablo_etiketi)
    if capa is None:
        return pd.DataFrame(columns=kolonlar)
    tablo_satir, _ = capa

    konum = None
    for satir in range(tablo_satir, tablo_satir + 6):
        for varyant in _IL_BASLIK_VARYANTLARI:
            hedef = normalize_label(varyant)
            for col in (1, 2):
                if normalize_label(ws.cell(row=satir, column=col).value) == hedef:
                    konum = (satir, col)
                    break
            if konum:
                break
        if konum:
            break
    if konum is None:
        return pd.DataFrame(columns=kolonlar)
    baslik_satir, il_sutun = konum
    deger_sutun = il_sutun + 1

    satirlar = []
    for satir_no, il_adi in _veri_satirlarini_gez(ws, baslik_satir, il_sutun):
        deger = parse_sayi(ws.cell(row=satir_no, column=deger_sutun).value)
        satirlar.append(
            {
                "il": il_adi,
                "il_kodu": il_kodu_bul(il_adi),
                "tarih_id": tarih_id,
                "lisans": lisans,
                deger_kolon_adi: deger,
            }
        )
    return pd.DataFrame(satirlar, columns=kolonlar)


def tablo2_uretim_kaynak_oku(ws: Worksheet, tarih_id: int) -> pd.DataFrame:
    """T2: ülke geneli, kaynak bazında lisanslı üretim (MWh; il yok)."""
    return tablo_kaynak_toplam_oku(ws, "Tablo 2", tarih_id, "uretim_mwh", "Lisanslı")


def tablo3_uretim_il_oku(ws: Worksheet, tarih_id: int) -> pd.DataFrame:
    """T3: il bazında toplam lisanslı üretim (MWh; kaynak kırılımı yok)."""
    return tablo_il_toplam_oku(ws, "Tablo 3", tarih_id, "uretim_mwh", "Lisanslı")


def tablo5_lisanssiz_uretim_kaynak_oku(ws: Worksheet, tarih_id: int) -> pd.DataFrame:
    """T5: ülke geneli, kaynak bazında lisanssız üretim (MWh; il yok)."""
    return tablo_kaynak_toplam_oku(ws, "Tablo 5", tarih_id, "uretim_mwh", "Lisanssız")


def tablo6_lisanssiz_uretim_il_oku(ws: Worksheet, tarih_id: int) -> pd.DataFrame:
    """T6: il bazında toplam lisanssız üretim (MWh; kaynak kırılımı yok)."""
    return tablo_il_toplam_oku(ws, "Tablo 6", tarih_id, "uretim_mwh", "Lisanssız")


def uretim_kaynak_birlestir(
    kurulu_il_kaynak_df: pd.DataFrame, uretim_kaynak_df: pd.DataFrame
) -> pd.DataFrame:
    """T1/T4 (il×kaynak kurulu güç) kaynak bazında toplanır, T2/T5 (ülke geneli
    üretim) ile 'kaynak' üzerinden birleştirilir. Sonuç ülke geneli, kaynak
    bazında (il YOK) bir DataFrame'dir — worker/kpi.py'nin ulusal KPI'ları
    (KPI-01..07) için kullanılabilir. İl×kaynak kesişiminde uretim_mwh yoktur
    (bkz. modül notu); il bazlı kurulu güç detayı ayrı kalır.
    """
    if kurulu_il_kaynak_df.empty:
        kurulu_kaynak = pd.DataFrame(
            columns=["kaynak", "yenilenebilir", "lisans", "tarih_id", "kurulu_guc_mw"]
        )
    else:
        kurulu_kaynak = kurulu_il_kaynak_df.groupby(
            ["kaynak", "yenilenebilir", "lisans", "tarih_id"], as_index=False
        )["kurulu_guc_mw"].sum(min_count=1)
    ortak = ["kaynak", "yenilenebilir", "lisans", "tarih_id"]
    return kurulu_kaynak.merge(uretim_kaynak_df, on=ortak, how="outer")


# ---------------------------------------------------------------------------
# Doğrulama yardımcıları (dokumanlar/05_kaynak_dosya_sozlesmesi.md — Doğrulama)
# ---------------------------------------------------------------------------


_TABLO_NUMARA_DESENI = re.compile(r"TABLO\s*(\d+)(?:\s*-\s*(\d+))?")


def _tablo_numaralarini_cikar(etiket: str) -> set[int]:
    """'Tablo 2-3' → {2, 3}; 'Tablo 11' → {11}. Sayfa adı gerçek dosyada birden
    fazla tabloyu aynı anda barındırabilir (ör. 'Tablo 2-3' hem Tablo 2'yi hem
    Tablo 3'ü içerir)."""
    eslesme = _TABLO_NUMARA_DESENI.search(normalize_label(etiket))
    if eslesme is None:
        return set()
    ilk = int(eslesme.group(1))
    if eslesme.group(2) is None:
        return {ilk}
    return set(range(ilk, int(eslesme.group(2)) + 1))


def sayfa_bul(wb: Workbook, tablo_no: int) -> Worksheet | None:
    """Sayfa adındaki tablo numarasına göre worksheet döner (ör. gerçek dosyada
    'Tablo 9-10' birleşik sayfası hem tablo_no=9 hem tablo_no=10 için eşleşir;
    sentetik test dosyasındaki ayrı 'Tablo 9'/'Tablo 10' sayfaları da aynı
    şekilde çalışır). Ay değiştikçe sayfa adlandırması farklılaşsa bile
    orkestrasyon (worker/pipeline.py) sabit sayfa adına bağımlı kalmaz."""
    for ad in wb.sheetnames:
        if tablo_no in _tablo_numaralarini_cikar(ad):
            return wb[ad]
    return None


def eksik_tablolari_bul(
    wb_sheet_names: list[str], gerekli_tablolar: list[str]
) -> list[str]:
    """13 tablo mevcut mu; eksikse batch reddi (sayfa adındaki tablo numarası bazlı)."""
    mevcut_numaralar: set[int] = set()
    for sayfa in wb_sheet_names:
        mevcut_numaralar |= _tablo_numaralarini_cikar(sayfa)
    eksik = []
    for tablo in gerekli_tablolar:
        gerekli_numaralar = _tablo_numaralarini_cikar(tablo)
        if not gerekli_numaralar <= mevcut_numaralar:
            eksik.append(tablo)
    return eksik


def mutabakat_kontrol(hesaplanan: float, resmi: float, tolerans: float = 0.005) -> bool:
    """İl toplamı ↔ 'TÜRKİYE' satırı ±%0,5 mutabakat kontrolü."""
    if resmi == 0:
        return hesaplanan == 0
    return abs(hesaplanan - resmi) / abs(resmi) <= tolerans
